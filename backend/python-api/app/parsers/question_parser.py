"""
문제 및 정답 데이터 파싱 서비스 (Gemini 2.0 Flash 기반)

모든 파일 형식을 Gemini API로 통합 처리
"""
import json
from typing import List, Dict, Any, Optional, Union, cast
from datetime import datetime, timezone
import google.generativeai as genai  # type: ignore
import os
from pathlib import Path
import base64
import logging
import re

from app.models.question import DifficultyLevel
from app.core.config import settings

logger = logging.getLogger(__name__)

# .env 자동 로드 (상위 디렉터리까지 탐색)
try:
    from dotenv import load_dotenv, find_dotenv
    load_dotenv(find_dotenv(usecwd=True), override=False)
except Exception:
    pass

# Poppler 경로 설정 (PDF→이미지 변환용): 협업 환경을 위한 동적 탐색
def _resolve_poppler_path() -> Optional[str]:
    """Poppler 경로를 환경에 맞게 동적으로 찾습니다.
    우선순위: 환경변수(POPPLER_PATH) > 설정(settings.POPPLER_PATH) > 로컬 번들 > 미설정(None)
    None을 반환하면 pdf2image가 시스템 PATH의 Poppler를 사용합니다.
    """
    try:
        # 1) 환경변수 우선
        path_from_env = os.getenv("POPPLER_PATH")
        if path_from_env and os.path.exists(path_from_env):
            return path_from_env

        # 2) settings에 정의된 경우
        path_from_settings = getattr(settings, "POPPLER_PATH", None)
        if path_from_settings and os.path.exists(path_from_settings):
            return path_from_settings

        # 3) 레포 내 번들된 경로 후보 (Windows 배포용)
        local_candidates = [
            Path(__file__).resolve().parent / "poppler-24.08.0" / "Library" / "bin",
            Path(__file__).resolve().parent / "poppler" / "bin",
            Path(__file__).resolve().parent.parent.parent / "poppler-windows" / "Library" / "bin",  # New installation location
        ]
        for candidate in local_candidates:
            if candidate.exists():
                return str(candidate)

        # 4) None을 반환하면 시스템 PATH 활용
        return None
    except Exception:
        return None

POPPLER_PATH = _resolve_poppler_path()


class QuestionParser:
    """gemini-2.0-flash-exp 기반 통합 파서"""
    
    def __init__(self, api_key: Optional[str] = None):
        """
        Args:
            api_key: Gemini API 키 (None인 경우 환경변수에서 가져옴)
        """
        # 키 우선순위: 직접 인자 > settings.GEMINI_API_KEY > 환경변수 GEMINI_API_KEY > 환경변수 GOOGLE_API_KEY
        self.api_key = (
            api_key
            or getattr(settings, "GEMINI_API_KEY", None)
            or os.getenv("GEMINI_API_KEY", "")
            or os.getenv("GOOGLE_API_KEY", "")
        )
        self.model: Optional[Any] = None  # Type annotation for model
        
        if self.api_key:
            genai.configure(api_key=self.api_key)  # type: ignore
            # Gemini 2.0 Flash 모델 사용 (정확한 모델명으로 수정)
            model_name = getattr(settings, 'GEMINI_MODEL_NAME', 'gemini-2.0-flash-exp')
            self.model = genai.GenerativeModel(model_name)  # type: ignore
            logger.info(f"Gemini 모델 초기화 완료: {model_name}")
        else:
            logger.warning("Gemini API 키가 설정되지 않았습니다.")
    
    def parse_any_file(self, file_path: str, content_type: str = "auto") -> Dict[str, Any]:
        """
        모든 파일 형식을 Gemini로 파싱 (분할 파싱 지원)
        
        Args:
            file_path: 파일 경로
            content_type: "questions", "answers", 또는 "auto" (자동 감지)
            
        Returns:
            파싱된 데이터
        """
        if not self.model:
            # 오프라인 폴백: 텍스트 파일일 때 간단 규칙 기반 파싱 수행
            file_extension = Path(file_path).suffix.lower()
            if file_extension in [".txt", ""]:
                try:
                    data = self._process_text_file_fallback(file_path, content_type)
                    return {"type": content_type if content_type != "auto" else "questions", "data": data}
                except Exception as e:
                    logger.error(f"오프라인 텍스트 폴백 실패: {e}")
            logger.error("Gemini API가 초기화되지 않았습니다.")
            return {"type": content_type, "data": [], "error": "Gemini API not initialized"}
        
        if not os.path.exists(file_path):
            logger.error(f"파일이 존재하지 않습니다: {file_path}")
            return {"type": content_type, "data": [], "error": "File not found"}
        
        # 파일 크기 확인
        file_size = os.path.getsize(file_path)
        logger.info(f"파일 크기: {file_size / (1024*1024):.2f} MB")
        
        # DB 스키마 정보
        db_schema = """
우리 데이터베이스 구조:

Question 테이블:
- question_number: 문제 번호 (정수)
- content: 문제 내용 (텍스트)
- description: 문제 설명/지문 (문자열 배열, 예: ["- 설명1", "- 설명2"])
- options: {"1": "선택지1", "2": "선택지2", ..., "5": "선택지5"}
- correct_answer: 정답 (문자열, 예: "3")
- subject: 과목명
- area_name: 영역이름
- difficulty: "하", "중", "상" 중 하나
- year: 연도 (정수)
"""
        
        try:
            # 파일 확장자에 따라 처리
            file_extension = Path(file_path).suffix.lower()
            
            if file_extension in ['.xlsx', '.xls']:
                all_data = self._process_excel_file_chunked(file_path, content_type, db_schema)
            elif file_extension == '.pdf':
                all_data = self._process_pdf_with_images(file_path, content_type, db_schema)
            else:
                all_data = self._process_text_file_chunked(file_path, content_type, db_schema)
            
            logger.info(f"분할 파싱 완료: {file_path}, 총 데이터 개수: {len(all_data) if isinstance(all_data, list) else 0}")
            return {"type": content_type if content_type != "auto" else "questions", "data": all_data}
                
        except Exception as e:
            logger.error(f"분할 파싱 오류 ({file_path}): {e}")
            return {"type": content_type, "data": [], "error": str(e)}
    
    def _generate_prompt(self, file_path: str, content_type: str, db_schema: str) -> str:
        """프롬프트 생성"""
        file_name = Path(file_path).name
        
        if content_type == "auto":
            return f"""다음 파일을 분석해주세요.

파일명: {file_name}

이 파일이 문제 데이터인지 정답 데이터인지 자동으로 판단하고,
아래 데이터베이스 구조에 맞게 JSON으로 변환해주세요.

{db_schema}

반환 형식:
{{
    "type": "questions" 또는 "answers",
    "data": [
        // 위 스키마에 맞는 객체들
    ]
}}

주의사항:
- 문제번호는 반드시 정수로 처리
- 선택지 번호는 "1", "2", "3", "4", "5" 문자열로
- 난이도: "하", "중", "상" 중 하나로 표시
- 연도는 파일명이나 내용에서 추출
- 없는 필드는 null로 표시
- JSON 형식으로만 응답해주세요"""
            
        elif content_type == "questions":
            return f"""이 파일은 시험 문제입니다.
            
{db_schema}

위 Question 스키마에 맞게 모든 문제를 JSON 배열로 변환해주세요.

예상되는 문제 형식:
번호. 문제내용
┌─────────────────────────────────────┐
│ 박스 (문제에 대한 추가설명/조건/지문)    │
└─────────────────────────────────────┘
1번 선택지 내용
2번 선택지 내용  
3번 선택지 내용
4번 선택지 내용
5번 선택지 내용

파싱 규칙:
- 문제번호 다음의 점(.) 뒤가 문제 내용(content)
- **박스 안의 내용이나 들여쓰기된 설명은 description 배열에 저장**
- 1번~5번 형태의 선택지는 options에 "1", "2", "3", "4", "5"로 저장
- 선택지가 ①②③④⑤로 되어있다면 "1", "2", "3", "4", "5"로 변환하세요

**박스 및 설명 추출 규칙:**
1. **박스 인식:** 다음과 같은 형태들을 찾으세요:
   - ┌─┐ └─┘ 또는 ╭─╮ ╰─╯ 형태의 선 박스
   - +---+ |   | +---+ 형태의 ASCII 박스
   - [박스 내용] 형태의 대괄호 박스
   - 「박스 내용」 형태의 코너 박스
   - 들여쓰기로 구분된 영역
   - "-" 또는 "*" 로 시작하는 목록 항목들

2. **박스 내용 추출:** 박스 안의 모든 텍스트를 description 배열에 포함:
   - "다음 조건을 만족하는..." 같은 조건문
   - "다음 그림/표를 보고..." 같은 지시문
   - 긴 설명 텍스트나 지문
   - 보기나 예시에 대한 설명

중요 제한사항:
- 모든 문제를 파싱하세요. 문제 수에 제한이 없습니다.

주의사항:
- 박스나 들여쓰기로 구분된 보충 설명은 description 필드에 배열로 저장하세요
- description은 문제를 풀기 위한 추가 정보나 조건들을 담습니다
- 예: ["- 몸에 널리 분포하며, 몸의 구조를 이룸", "- 세포나 기관 사이 틈을 메우고, 기관을 지지·보호함"]
- JSON 형식으로만 응답해주세요"""
        else:  # answers
            return f"""이 파일은 시험 정답 데이터입니다.

{db_schema}

위 스키마의 정답 관련 필드들(question_number, correct_answer, subject, area_name, difficulty)을 
JSON 배열로 변환해주세요.

JSON 형식으로만 응답해주세요."""
    


    def _process_excel_file_chunked(self, file_path: str, content_type: str, db_schema: str) -> List[Dict[str, Any]]:
        """Excel 파일 분할 처리 (openpyxl 사용)"""
        try:
            from openpyxl import load_workbook
        except ImportError:
            logger.error("openpyxl 라이브러리가 설치되지 않았습니다.")
            raise ImportError("openpyxl이 필요합니다. pip install openpyxl로 설치하세요.")
        
        all_data = []
        
        try:
            workbook = load_workbook(file_path, read_only=True)
            logger.info(f"Excel 파일 시트 목록: {workbook.sheetnames}")
            
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                logger.info(f"시트 '{sheet_name}' 처리 중...")
                
                # 시트 데이터를 텍스트로 변환
                sheet_data = []
                row_count = 0
                
                for row in worksheet.iter_rows(values_only=True):
                    if row_count >= 100:  # 시트당 최대 100행
                        break
                    if any(cell for cell in row):  # 빈 행이 아닌 경우만
                        sheet_data.append('\t'.join([str(cell) if cell is not None else '' for cell in row]))
                        row_count += 1
                
                if sheet_data:
                    sheet_text = '\n'.join(sheet_data)
                    
                    # Gemini로 구조화 요청
                    prompt = self._generate_prompt(f"{file_path} (시트: {sheet_name})", content_type, db_schema)
                    structured_prompt = f"""
다음은 Excel 시트 '{sheet_name}'의 데이터입니다.
이 데이터를 분석하여 구조화된 JSON으로 변환해주세요.

{prompt}

Excel 데이터:
{sheet_text}

중요: 22번 문제까지만 처리하세요.
"""
                    
                    try:
                        if self.model is None:
                            logger.error("Gemini 모델이 초기화되지 않았습니다.")
                            continue
                        
                        response = self.model.generate_content([structured_prompt])  # type: ignore
                        if response and response.text:
                            sheet_results = self._parse_gemini_response(response.text, content_type)
                            sheet_data_parsed = sheet_results.get("data", [])
                            
                            # 기본 전처리
                            sheet_data_parsed = sheet_results.get("data", [])
                            
                            if sheet_data_parsed:
                                # year 보정: Gemini가 year를 못 뽑았거나 0/None이면 시트명에서 추출
                                year_in_sheet = None
                                match = re.search(r'(20\d{2})', sheet_name)
                                if match:
                                    year_in_sheet = int(match.group(1))
                                else:
                                    year_in_sheet = 2020  # 기본값
                                for item in sheet_data_parsed:
                                    if not item.get('year') or item.get('year') in [0, None, '']:
                                        item['year'] = year_in_sheet
                                logger.info(f"시트 '{sheet_name}': {len(sheet_data_parsed)}개 항목 파싱 성공 (연도 보정: {year_in_sheet})")
                                all_data.extend(sheet_data_parsed)
                            else:
                                logger.warning(f"시트 '{sheet_name}': 파싱된 데이터 없음")
                    except Exception as e:
                        logger.error(f"시트 '{sheet_name}' 파싱 실패: {e}")
                        continue
                        
                # 22개 달성하면 중단
                if len(all_data) >= 22:
                    all_data = all_data[:22]
                    break
            
            workbook.close()
            logger.info(f"Excel 파일 처리 완료: 총 {len(all_data)}개 항목")
            return all_data
            
        except Exception as e:
            logger.error(f"Excel 파일 읽기 오류: {e}")
            raise
    
    def _process_pdf_with_images(self, file_path: str, content_type: str, db_schema: str) -> List[Dict[str, Any]]:
        """PDF 파일을 이미지로 변환하여 Gemini로 처리 (PyPDF2 사용 안함)"""
        try:
            from pdf2image import convert_from_path
        except ImportError:
            logger.error("pdf2image 라이브러리가 설치되지 않았습니다.")
            raise ImportError("pdf2image가 필요합니다. pip install pdf2image로 설치하세요.")
        
        all_questions = []
        
        try:
            # PDF를 이미지로 변환 (pdf2image의 역할)
            logger.info("PDF를 이미지로 변환 중...")
            
            # POPPLER_PATH가 None인 경우 인수에서 제외
            convert_kwargs: Dict[str, Any] = {
                "dpi": 200  # 고품질 이미지
            }
            if POPPLER_PATH is not None:
                convert_kwargs["poppler_path"] = POPPLER_PATH
            
            page_images = convert_from_path(file_path, **convert_kwargs)
            
            logger.info(f"총 {len(page_images)}개 페이지 이미지 생성됨")
            
            # Gemini가 이미지를 읽고 이해하고 구조화하는 프롬프트
            gemini_prompt = f"""
이 이미지를 분석하여 시험 문제를 찾아 구조화된 JSON으로 변환해주세요.

{db_schema}

예상되는 문제 형식:
번호. 문제내용
┌─────────────────────────────────────┐
│ 박스 (문제에 대한 추가설명/조건/지문)    │
└─────────────────────────────────────┘
1번 선택지 내용
2번 선택지 내용  
3번 선택지 내용
4번 선택지 내용
5번 선택지 내용

**박스 감지 및 추출 규칙 (매우 중요!):**
1. **박스 형태 인식:** 다음과 같은 박스 형태들을 모두 찾아내세요:
   - ┌─┐ └─┘ 형태의 선으로 그려진 박스
   - ╭─╮ ╰─╯ 형태의 둥근 박스
   - +--+ |  | +--+ 형태의 ASCII 박스
   - 실선으로 둘러싸인 사각형 박스
   - 점선으로 둘러싸인 박스
   - 배경이 회색/음영으로 처리된 박스
   - 테두리가 굵게 표시된 영역

2. **박스 내용 추출:** 박스 안의 모든 텍스트를 description 배열에 포함하세요:
   - 박스 안의 설명문
   - 조건문 ("다음 조건을 만족하는...")
   - 지문 (긴 설명 텍스트)
   - 표나 그림에 대한 설명
   - 들여쓰기된 부가 정보

3. **박스 위치 관계:** 박스가 문제번호 바로 다음에 나오거나, 문제 내용 중간에 삽입된 경우 모두 찾아내세요.

위 스키마에 맞게 각 문제를 추출하세요:
- question_number: 문제 번호
- content: 문제 내용 (번호 다음 점 뒤의 내용, 박스 내용은 제외)
- description: **박스 안의 모든 내용을 배열로 저장** (예: ["조건: 다음을 만족하는 경우", "- 추가 설명 내용"])
- options: 선택지 (1번~5번 또는 ①②③④⑤ → "1","2","3","4","5")
- subject: 과목명 (이미지에서 추출 가능하면)
- area_name: 영역명 (이미지에서 추출 가능하면)
- difficulty: "하", "중", "상" 중 하나
- year: 연도 (이미지에서 추출 가능하면)

**중요:** 박스가 보이지 않더라도 들여쓰기나 구분된 영역의 설명문이 있다면 description에 포함하세요.

JSON 배열로만 응답하세요.
"""
            
            # 각 페이지 이미지를 Gemini가 읽고 이해하고 구조화
            for page_num, page_image in enumerate(page_images, 1):
                
                logger.info(f"페이지 {page_num} 이미지를 Gemini로 분석 중...")
                
                try:
                    # Gemini가 이미지를 읽고 이해하고 구조화
                    if self.model is None:
                        logger.error("Gemini 모델이 초기화되지 않았습니다.")
                        continue
                    
                    response = self.model.generate_content([gemini_prompt, page_image])  # type: ignore
                    
                    if response and response.text:
                        try:
                            # Gemini 응답 파싱
                            page_result = self._parse_gemini_response(response.text, content_type)
                            page_questions = page_result.get("data", [])
                            
                            if page_questions:
                                logger.info(f"페이지 {page_num}: {len(page_questions)}개 문제 추출")
                                all_questions.extend(page_questions)
                            else:
                                logger.info(f"페이지 {page_num}: 추출된 문제 없음")
                                
                        except Exception as e:
                            logger.error(f"페이지 {page_num} 파싱 실패: {e}")
                            continue
                    else:
                        logger.warning(f"페이지 {page_num}: Gemini 응답 없음")
                    
                except Exception as e:
                    logger.error(f"페이지 {page_num} Gemini 분석 실패: {e}")
                    continue
            
            logger.info(f"PDF 이미지 분석 완료: 총 {len(all_questions)}개 문제")
            return all_questions
            
        except Exception as e:
            logger.error(f"PDF 이미지 처리 실패: {e}")
            raise  # PyPDF2 폴백 제거, 에러 발생시 예외 발생



    def _process_text_file_chunked(self, file_path: str, content_type: str, db_schema: str) -> List[Dict[str, Any]]:
        """텍스트 파일 분할 처리"""
        encodings = ['utf-8', 'cp949', 'euc-kr', 'latin-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding) as f:
                    content = f.read()
                break
            except UnicodeDecodeError:
                continue
        else:
            raise ValueError(f"파일 인코딩을 감지할 수 없습니다: {file_path}")
        
        return self._process_text_chunks(content, content_type, db_schema)

    def _process_text_chunks(self, content: str, content_type: str, db_schema: str) -> List[Dict[str, Any]]:
        """텍스트 내용을 청크로 분할 처리"""
        all_data = []
        chunk_size = 15000  # 문자 단위 청크 크기
        
        # 텍스트가 작은 경우 한번에 처리
        if len(content) <= chunk_size:
            try:
                if self.model is None:
                    logger.error("Gemini 모델이 초기화되지 않았습니다.")
                    return []
                
                prompt = self._generate_prompt("text_content", content_type, db_schema)
                response = self.model.generate_content([prompt, f"텍스트 내용:\n{content}"])  # type: ignore
                result = self._parse_gemini_response(response.text, content_type).get("data", [])
                # 22번 제한 적용
                result = [item for item in result if item.get('question_number', 0) <= 22][:22]
                return result
            except Exception as e:
                logger.error(f"텍스트 처리 오류: {e}")
                return []
        
        # 큰 텍스트는 청크로 분할
        for i in range(0, len(content), chunk_size):
            chunk = content[i:i + chunk_size]
            
            try:
                if self.model is None:
                    logger.error("Gemini 모델이 초기화되지 않았습니다.")
                    continue
                
                prompt = self._generate_prompt("text_content", content_type, db_schema)
                response = self.model.generate_content([prompt, f"텍스트 청크:\n{chunk}"])  # type: ignore
                chunk_data = self._parse_gemini_response(response.text, content_type).get("data", [])
                
                all_data.extend(chunk_data)
                logger.info(f"텍스트 청크 처리 완료: {len(chunk_data)}개 데이터")
                    
            except Exception as e:
                logger.warning(f"텍스트 청크 처리 실패: {e}")
                continue
        
        return all_data

    def _process_text_file_fallback(self, file_path: str, content_type: str) -> List[Dict[str, Any]]:
        """Gemini 없이 텍스트 파일을 간단 규칙으로 파싱하는 폴백.
        - 형식 예시:
          1. 문제 내용
          ① 보기1  ② 보기2  ③ 보기3  ④ 보기4  ⑤ 보기5
        - 정답/과목/영역/연도는 미설정
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='cp949') as f:
                content = f.read()

        import re
        lines = [l.strip() for l in content.splitlines() if l.strip()]
        questions: List[Dict[str, Any]] = []
        current: Dict[str, Any] = {}

        def push_current():
            if current.get('content'):
                questions.append({
                    'question_number': current.get('question_number', len(questions) + 1),
                    'content': current['content'],
                    'description': current.get('description', []),
                    'options': current.get('options', {}),
                    'subject': current.get('subject', ''),
                    'area_name': current.get('area_name', ''),
                    'difficulty': current.get('difficulty', '중'),
                    'year': current.get('year')
                })

        num_pat = re.compile(r'^(\d{1,2})\s*[\.|\)]\s*(.+)$')
        # 옵션 패턴: ①②③④⑤ 또는 숫자 기반
        opt_pat = re.compile(r'[①②③④⑤]|\b1\b|\b2\b|\b3\b|\b4\b|\b5\b')

        for line in lines:
            m = num_pat.match(line)
            if m:
                # 새 문제 시작
                if current:
                    push_current()
                current = {
                    'question_number': int(m.group(1)),
                    'content': m.group(2).strip(),
                    'options': {}
                }
                continue

            if opt_pat.search(line):
                # 옵션 라인 처리
                # ①~⑤를 1~5로 치환 후 분할 시도
                normalized = line.replace('①', '1').replace('②', '2').replace('③', '3').replace('④', '4').replace('⑤', '5')
                # 구분자 기준 분할
                parts = re.split(r'\s{2,}|\s\s+|\t|\s-\s|\s\|\s', normalized)
                # 보수적으로 공백으로도 분할
                if len(parts) < 2:
                    parts = re.split(r'\s{1,}', normalized)
                # 라벨과 텍스트를 재구성
                label = None
                buffer: Dict[str, str] = {}
                for token in parts:
                    token = token.strip()
                    if not token:
                        continue
                    if token in {'1', '2', '3', '4', '5'}:
                        label = token
                        buffer[label] = ''
                    else:
                        if label is None:
                            continue
                        sep = '' if buffer[label] == '' else ' '
                        buffer[label] = f"{buffer[label]}{sep}{token}"
                if current:
                    current['options'] = {k: v.strip() for k, v in buffer.items() if v.strip()}
                continue

            # 기타 라인은 본문에 이어붙임
            if current:
                current['content'] = (current.get('content', '') + ' ' + line).strip()

        if current:
            push_current()

        # answers 타입 요청이면 옵션 없이 스킵(오프라인 폴백에서는 미처리)
        if content_type == 'answers':
            return []
        return questions


    
    def _clean_json_text(self, text: str) -> str:
        """JSON 텍스트에서 주석 및 불필요한 요소 제거"""
        lines = text.split('\n')
        cleaned_lines = []
        
        for line in lines:
            # // 주석 제거
            if '//' in line:
                # JSON 문자열 내부의 //는 보존
                in_string = False
                escaped = False
                cleaned_line = ""
                
                for i, char in enumerate(line):
                    if escaped:
                        cleaned_line += char
                        escaped = False
                        continue
                    
                    if char == '\\':
                        escaped = True
                        cleaned_line += char
                        continue
                    
                    if char == '"' and not escaped:
                        in_string = not in_string
                        cleaned_line += char
                        continue
                    
                    if not in_string and char == '/' and i + 1 < len(line) and line[i + 1] == '/':
                        # 주석 시작, 나머지 줄 무시
                        break
                    
                    cleaned_line += char
                
                line = cleaned_line
            
            # /* */ 주석 제거 (단순 버전)
            while '/*' in line and '*/' in line:
                start = line.find('/*')
                end = line.find('*/', start) + 2
                line = line[:start] + line[end:]
            
            # 빈 줄이 아니면 추가
            if line.strip():
                cleaned_lines.append(line)
        
        return '\n'.join(cleaned_lines)
    
    def _aggressive_json_clean(self, text: str) -> str:
        """더 적극적인 JSON 정리"""
        import re
        
        # 마지막 } 또는 ] 이후의 모든 텍스트 제거
        text = text.strip()
        
        # JSON 배열인지 객체인지 확인
        if text.startswith('['):
            # 마지막 ]의 위치 찾기
            last_bracket = text.rfind(']')
            if last_bracket != -1:
                text = text[:last_bracket + 1]
        elif text.startswith('{'):
            # 마지막 }의 위치 찾기
            last_brace = text.rfind('}')
            if last_brace != -1:
                text = text[:last_brace + 1]
        
        # 여러 개의 공백을 하나로 압축
        text = re.sub(r'\s+', ' ', text)
        
        # 줄바꿈 정리
        text = text.replace('\n', ' ').replace('\r', '')
        
        # 문자열 외부의 주석 제거 (간단한 버전)
        text = re.sub(r'//[^\n\r]*', '', text)
        
        return text.strip()
    
    def _parse_gemini_response(self, response_text: str, content_type: str) -> Dict[str, Any]:
        """Gemini 응답 파싱"""
        text = response_text.strip()
        
        # JSON 추출
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0]
        elif "```" in text:
            json_parts = text.split("```")
            for part in json_parts:
                if part.strip().startswith('{') or part.strip().startswith('['):
                    text = part
                    break
        
        # JSON 주석 제거 (Gemini가 종종 주석을 포함함)
        text = self._clean_json_text(text)
        
        # JSON 파싱 시도
        try:
            result = json.loads(text.strip())
        except json.JSONDecodeError as e:
            logger.error(f"JSON 파싱 오류: {e}")
            logger.error(f"정리된 텍스트: {text[:500]}...")
            
            # 두 번째 시도: 더 적극적인 정리
            cleaned_text = self._aggressive_json_clean(text)
            try:
                result = json.loads(cleaned_text)
                logger.info("두 번째 시도로 JSON 파싱 성공")
            except json.JSONDecodeError as e2:
                logger.error(f"두 번째 JSON 파싱도 실패: {e2}")
                logger.error(f"최종 텍스트: {cleaned_text[:500]}...")
            raise ValueError(f"Invalid JSON response: {e}")
        except Exception as e:
            logger.error(f"예상치 못한 오류: {e}")
            raise ValueError(f"Unexpected error in JSON parsing: {e}")
        
        # 자동 감지 모드인 경우
        if content_type == "auto" and isinstance(result, dict) and "type" in result:
            data = result.get("data", [])
            return {
                "type": result["type"],
                "data": data
            }
        else:
            # 지정된 타입인 경우
            data = result if isinstance(result, list) else result.get("data", [])
            return {
                "type": content_type,
                "data": data
            }
    
    def match_questions_with_answers(
        self, 
        questions: List[Dict[str, Any]], 
        answers: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """
        문제와 정답 매칭 (부분 매칭 지원)
        
        정답이 부족한 경우에도 매칭 가능한 범위까지만 완전한 데이터를 반환
        """
        # 입력 데이터에서 정답을 문제번호로 인덱싱
        answer_map = {}
        for ans in answers:
            q_num = ans.get("question_number")
            if q_num is not None:
                answer_map[str(q_num)] = ans
        
        matched_data = []
        skipped_count = 0
        
        # 정답이 있는 문제번호 범위 확인
        if answer_map:
            available_answer_numbers = set(answer_map.keys())
            logger.info(f"사용 가능한 정답: {len(available_answer_numbers)}개 문제 ({min(available_answer_numbers) if available_answer_numbers else 'N/A'} ~ {max(available_answer_numbers) if available_answer_numbers else 'N/A'}번)")
        else:
            logger.warning("정답 데이터가 없습니다.")
            available_answer_numbers = set()
        
        for question in questions:
            q_num = question.get("question_number")
            if q_num is None:
                logger.warning(f"문제번호가 없는 문제: {question.get('content', '')[:50]}...")
                skipped_count += 1
                continue
                
            q_num_str = str(q_num)
            
            # 정답이 있는 경우만 처리 (완전한 데이터만 반환)
            if q_num_str in answer_map:
                answer_data = answer_map[q_num_str]
                
                # 문제와 정답 데이터 병합
                matched_item = {
                    **question,
                    "correct_answer": answer_data.get("correct_answer") or answer_data.get("answer", ""),
                    "subject": answer_data.get("subject", question.get("subject", "")),
                    "area_name": answer_data.get("area_name", question.get("area_name", "")),
                    "difficulty": answer_data.get("difficulty", question.get("difficulty", "중")),
                    "year": answer_data.get("year", question.get("year")),
                    "answer_source": "matched"
                }
                
                # 필수 필드 검증 (완전한 데이터만 포함)
                if self._is_complete_question_data(matched_item):
                    matched_data.append(matched_item)
                else:
                    logger.warning(f"불완전한 데이터로 인해 문제 {q_num} 제외")
                    skipped_count += 1
            else:
                # 정답이 없는 문제는 제외 (부분 매칭 정책)
                logger.debug(f"정답이 없어서 제외된 문제: {q_num}")
                skipped_count += 1
        
        # 매칭 결과 로깅
        total_questions = len(questions)
        matched_count = len(matched_data)
        
        logger.info(f"매칭 완료: 전체 {total_questions}개 문제 중 {matched_count}개 완전 매칭")
        logger.info(f"매칭률: {(matched_count/total_questions*100):.1f}%")
        
        if skipped_count > 0:
            logger.info(f"제외된 문제: {skipped_count}개 (정답 없음 또는 불완전한 데이터)")
        
        return matched_data
    
    def _is_complete_question_data(self, question_data: Dict[str, Any]) -> bool:
        """
        문제 데이터가 완전한지 검증
        
        Args:
            question_data: 검증할 문제 데이터
            
        Returns:
            bool: 완전한 데이터 여부
        """
        required_fields = ["question_number", "content", "correct_answer"]
        
        for field in required_fields:
            value = question_data.get(field)
            if value is None or (isinstance(value, str) and not value.strip()):
                return False
        
        # 선택지가 있는 경우 검증
        options = question_data.get("options", {})
        if options and len(options) < 2:
            return False
        
        return True
    
    def convert_to_db_format(self, matched_data: List[Dict[str, Any]]) -> Dict[str, List[Dict[str, Any]]]:
        """
        매칭된 데이터를 데이터베이스 테이블 형식으로 변환
        """
        questions = []
        answer_options = []
        correct_answers = []
        
        current_time = datetime.now(timezone.utc).isoformat()
        
        for idx, item in enumerate(matched_data, 1):
            question_id = idx
            
            # Question 레코드
            question_record = {
                "id": question_id,
                "content": item.get("content", ""),
                "description": item.get("description", None),
                "difficulty": item.get("difficulty", "중"),
                "subject": item.get("subject", ""),
                "area_name": item.get("area_name", ""),
                "is_active": True,
                "question_metadata": {
                    "question_number": item.get("question_number"),
                    "year": item.get("year"),
                    "answer_source": item.get("answer_source", ""),
                },
                "created_at": current_time,
                "updated_at": current_time
            }
            
            # 간단 버전 필드 추가
            options = item.get("options", {})
            if options:
                question_record["choices"] = list(options.values())
                question_record["correct_answer"] = item.get("correct_answer", "")
            
            questions.append(question_record)
            
            # AnswerOption 레코드들
            for option_label, option_text in options.items():
                option_record = {
                    "question_id": question_id,
                    "option_text": option_text,
                    "option_label": option_label,
                    "display_order": int(option_label) if option_label.isdigit() else 0,
                    "created_at": current_time,
                    "updated_at": current_time
                }
                answer_options.append(option_record)
            
            # CorrectAnswer 레코드
            correct_answer = item.get("correct_answer")
            if correct_answer:
                correct_answer_record = {
                    "question_id": question_id,
                    "answer_text": correct_answer,
                    "created_at": current_time,
                    "updated_at": current_time
                }
                correct_answers.append(correct_answer_record)
        
        logger.info(f"DB 형식 변환 완료: Questions({len(questions)}), Options({len(answer_options)}), Answers({len(correct_answers)})")
        
        return {
            "questions": questions,
            "answer_options": answer_options,
            "correct_answers": correct_answers
        }


# 싱글톤 인스턴스
question_parser = QuestionParser()